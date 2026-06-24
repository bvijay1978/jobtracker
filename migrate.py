"""Import a job-tracker spreadsheet into jobs.db (SQLite).

This is the importer for the specific .xlsx layout this project grew out of: a
workbook with three overlapping data sheets (Pipeline, Applied, Not Applied) and
the 13 columns listed in COL_TO_FIELD. It's optional — a fresh install starts
from an empty database (or `python seed.py` for demo data). Point it at your own
file with `--xlsx PATH` or the JOBTRACKER_XLSX environment variable.

Pipeline is the maintained master; Applied/Not Applied are mostly filtered — and
sometimes staler — views, but they occasionally hold rows the master is missing.
We therefore merge all three:

  * Dedupe key  : numeric job-id parsed from the link (the true unique key),
                  falling back to normalised title|company for rows with no URL.
  * Precedence  : sheets are processed Pipeline -> Applied -> Not Applied. The
                  first sheet to supply a field wins; later sheets only fill
                  blanks. Because Pipeline always carries a real status, its
                  status/applied-date/outcome are never clobbered by a staler
                  'Not Applied' view.
  * Links       : a cell displaying "View advert" stores the real URL as an
                  Excel hyperlink — we recover the target.
  * Status      : the literal 'Not Applied' (used as a status on that sheet) is
                  normalised to 'Found', matching the canonical vocabulary.

Re-runnable: pass --reset to rebuild jobs.db from scratch.
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

import config
import db
from normalize import clean, norm_date, norm_status, parse_job_id, tc_key

DATA_SHEETS = ["Pipeline", "Applied", "Not Applied"]

# Spreadsheet column order (A..M) -> our field names.
COL_TO_FIELD = [
    "date_found",   # A  Date Found
    "title",        # B  Job Title
    "company",      # C  Company / Recruiter
    "type",         # D  Type
    "rate",         # E  Rate / Salary
    "location",     # F  Location & Pattern
    "posted",       # G  Posted
    "link",         # H  Job URL
    "fit_notes",    # I  Fit Notes
    "status",       # J  Status
    "cv_version",   # K  CV Version Used
    "date_applied", # L  Date Applied
    "outcome",      # M  Outcome / Next Step
]
DATE_FIELDS = {"date_found", "posted", "date_applied"}


def resolve_link(cell) -> str | None:
    """Prefer the hyperlink target; treat bare 'View advert' text as no link."""
    target = cell.hyperlink.target if cell.hyperlink else None
    if target:
        return target.strip()
    text = clean(cell.value)
    if text and text.startswith("http"):
        return text
    return None  # e.g. "View advert" with no embedded hyperlink


def read_rows(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sheet in DATA_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            cells = [ws.cell(r, c) for c in range(1, 14)]
            values = [c.value for c in cells]
            if all(v is None for v in values):
                continue
            rec: dict = {}
            for idx, field in enumerate(COL_TO_FIELD):
                if field == "link":
                    rec[field] = resolve_link(cells[idx])
                elif field == "status":
                    rec[field] = norm_status(values[idx])
                elif field in DATE_FIELDS:
                    rec[field] = norm_date(values[idx])
                else:
                    rec[field] = clean(values[idx])
            rec["source_job_id"] = parse_job_id(rec["link"])
            if not rec.get("title"):
                continue  # skip stray/blank rows
            yield sheet, rec


def merge_into(existing: dict, incoming: dict) -> None:
    """Fill only blank fields on `existing` from `incoming` (first sheet wins)."""
    for field in (*db.FIELDS, "source_job_id"):
        if not existing.get(field) and incoming.get(field):
            existing[field] = incoming[field]


def build_unique_records(xlsx_path: Path):
    records: list[dict] = []
    by_jobid: dict[str, dict] = {}
    by_tc: dict[str, dict] = {}
    stats = {"read": 0, "merged": 0}

    for _sheet, rec in read_rows(xlsx_path):
        stats["read"] += 1
        jid = rec.get("source_job_id")
        tc = tc_key(rec.get("title"), rec.get("company"))
        existing = by_jobid.get(jid) if jid else None
        if existing is None:
            # Fall back to title|company, but never conflate two rows that carry
            # *different* job-ids — those are genuinely different postings.
            cand = by_tc.get(tc)
            if cand is not None:
                cand_jid = cand.get("source_job_id")
                if not jid or not cand_jid or cand_jid == jid:
                    existing = cand
        if existing is not None:
            merge_into(existing, rec)
            stats["merged"] += 1
        else:
            records.append(rec)
            existing = rec
        # register both keys so future rows collapse onto this record
        if jid:
            by_jobid.setdefault(jid, existing)
        by_tc.setdefault(tc, existing)
    return records, stats


def main() -> None:
    ap = argparse.ArgumentParser(description="Import a tracker spreadsheet into jobs.db")
    ap.add_argument("--xlsx", type=Path, default=config.XLSX_PATH,
                    help="Source spreadsheet (or set JOBTRACKER_XLSX)")
    ap.add_argument("--db", type=Path, default=db.DB_PATH, help="Target SQLite file")
    ap.add_argument("--reset", action="store_true", help="Drop and rebuild the jobs table first")
    args = ap.parse_args()

    if args.xlsx is None:
        raise SystemExit(
            "No spreadsheet specified. Pass --xlsx PATH or set JOBTRACKER_XLSX in "
            ".env. (To start empty, just run the app; for demo data run "
            "`python seed.py`.)"
        )
    if not args.xlsx.exists():
        raise SystemExit(f"Spreadsheet not found: {args.xlsx}")

    if args.reset and Path(args.db).exists():
        with db.connect(args.db) as conn:
            conn.execute("DROP TABLE IF EXISTS jobs")
    db.init_db(args.db)

    records, stats = build_unique_records(args.xlsx)

    with db.connect(args.db) as conn:
        existing_count = conn.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
        if existing_count and not args.reset:
            raise SystemExit(
                f"jobs.db already holds {existing_count} rows. "
                "Re-run with --reset to rebuild, or migrate into an empty DB."
            )
        for rec in records:
            db.insert_job(conn, rec)
        conn.commit()
        total = conn.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
        by_status = conn.execute(
            "SELECT status, COUNT(*) c FROM jobs GROUP BY status ORDER BY c DESC"
        ).fetchall()

    print(f"Read {stats['read']} sheet rows -> merged {stats['merged']} duplicates "
          f"-> {len(records)} unique jobs inserted.")
    print(f"jobs.db now holds {total} jobs:")
    for row in by_status:
        print(f"  {row['status'] or '(none)':<14} {row['c']}")


if __name__ == "__main__":
    main()
